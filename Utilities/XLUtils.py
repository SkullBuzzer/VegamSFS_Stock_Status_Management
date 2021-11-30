import openpyxl


def getRowCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return sheet.max_row


def getColumnCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return sheet.max_column


def readData(file, SheetName, rows, columns):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return sheet.cell(row=rows, column=columns).value


def writeData(file, SheetName, rows, columns, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    sheet.cell(row=rows, column=columns).value = data
    workbook.save(file)
